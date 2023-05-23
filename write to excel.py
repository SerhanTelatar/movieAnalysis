import requests
from bs4 import BeautifulSoup as bs
import openpyxl


def intBoxOfficeMaker(boxOfficeStr):
    boxOfficeStr = boxOfficeStr.replace("$", "").replace(",", "").replace("(estimated)", "").strip()
    boxOfficeInt = int(boxOfficeStr)
    return boxOfficeInt

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3'
    # bu kısım değişebilir
}

URL = "https://www.imdb.com/list/ls056515174/?st_dt=&mode=simple&page=1&sort=user_rating,desc"

page = requests.get(URL)

content = bs(page.content, 'html.parser')

wb = openpyxl.Workbook()
ws = wb.active

row_num = 2

ws.cell(row=1, column=1, value="Name")
ws.cell(row=1, column=2, value="Year")
ws.cell(row=1, column=3, value="Rate")
ws.cell(row=1, column=4, value="Box Office")
ws.cell(row=1, column=5, value="Link")

for item in content.find_all('div', {'class': 'lister-item mode-simple'}):
    link = "https://www.imdb.com/" + item.find('div', class_='lister-item-image').find('a')['href'] + "?ref_=ttls_li_tt"

    name = item.find('span', class_='lister-item-header').find('a').text.strip()
    year = int(item.find('span', class_='lister-item-year').text[1:5])
    rate = float(item.find('strong').text.strip())

    if 2000 < year:
        budgetPage = requests.get(link, headers=headers)
        budgetPageContent = bs(budgetPage.content, 'html.parser')
        try:
            boxOffice = budgetPageContent.find(attrs={"data-testid": "title-boxoffice-cumulativeworldwidegross"}).find('span',
                                                                                                                    class_='ipc-metadata-list-item__list-content-item').text
            boxOffice_int = intBoxOfficeMaker(boxOffice)


            ws.cell(row=row_num, column=1, value=name)
            ws.cell(row=row_num, column=2, value=year)
            ws.cell(row=row_num, column=3, value=rate)
            ws.cell(row=row_num, column=4, value=boxOffice_int)
            ws.cell(row=row_num, column=5, value=link)

            row_num += 1
        except:
            continue

wb.save("movie_data.xlsx")

